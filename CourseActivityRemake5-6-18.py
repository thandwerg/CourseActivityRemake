
#creates a dictionary of names and urls
#improve by reading class name from original spreadsheet

#creates initial data structure from downloaded apex report
import openpyxl, shelve, datetime
from openpyxl.chart import BarChart, Series, Reference, LineChart
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart.axis import DateAxis
wb= openpyxl.load_workbook('C:\\Programs\\Activity\\report.xlsx')
sheet=wb.active
studentdict={}

for x in range(1,sheet.max_row):
    name= sheet.cell(row=x, column=19).value
    if name not in list(studentdict.keys()):
        studentdict[name]={}
    enrollmentid= sheet.cell(row=x, column=13).value
    subject= sheet.cell(row=x, column=9).value
    urllist=studentdict[name]
    urllist['https://reports-prd.apexvs.com/ApexUI/Reports/Student/courseActivityScoreReport.aspx?enrollmentID=%s' %(enrollmentid)]={'coursename' : subject, 'datecount': {}}
    



#selenium stuff
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os, time, sys

#login and such
os.chdir('C:\\Programs')
browser = webdriver.Firefox()
browser.get('http://apexvs.com')
elementz = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.NAME, 'ctl00$ContentPlaceHolder1$loginUsernameTextBox')))
login = browser.find_element_by_name('ctl00$ContentPlaceHolder1$loginUsernameTextBox')
login.send_keys('login')
passwd= browser.find_element_by_name('ctl00$ContentPlaceHolder1$passwordTextBox')
passwd.send_keys('pass')
passwd.submit()
time.sleep(3) 

#scrapes the web data
for student in list(studentdict.keys()): #names

    os.chdir('C:\\Programs')
    a=studentdict[student]
    urllist= list(a.keys())
    try:
        
        for y in urllist:  #urls
        
            print(y)
            browser.get(y)
            cell = []
            celltext = []
            
            for x in range(100):   #populates the cell list with all completed assignments
                element = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.ID, 'reportGrid_cell_0_4')))
                cell.append(browser.find_elements_by_id('reportGrid_cell_%s_4' % x))
            for unit in cell:  #converts webelement to date
                for point in unit:
                    if len(point.text)>1:
                        celltext.append(point.text)
            b=a[y] #points to the value for the url key - should give coursename and datecount dict as output
            c=b['datecount']
            for apexdate in celltext: #creates a dictionary with key=date and value= number of assignments on that date

                middate= apexdate.split() #added
                            
                if 'Jan' in apexdate:
                    middate[1]='01'
                elif 'Feb' in apexdate:
                    middate[1]='02'
                elif 'Mar' in apexdate:
                    middate[1]='03'
                elif 'Apr' in apexdate:
                    middate[1]='04'
                elif 'May' in apexdate:
                    middate[1]='05'
                elif 'Jun' in apexdate:
                    middate[1]='06'
                elif 'Jul' in apexdate:
                    middate[1]='07'
                elif 'Aug' in apexdate:
                    middate[1]='08'
                elif 'Sep' in apexdate:
                    middate[1]='09'
                elif 'Oct' in apexdate:
                    middate[1]='10'
                elif 'Nov' in apexdate:
                    middate[1]='11'
                elif 'Dec' in apexdate:
                    middate[1]='12'


                newdate= "%s/%s/%s" % (middate[1],middate[0],middate[2])
                if newdate not in list(c.keys()):
                    c[newdate]=1
                else:
                    c[newdate]+=1
    except Exception as err:
        print('there was an error: %s' % (err))    
  
#deposits it in Excel

                
wb = openpyxl.Workbook()



#populate the worksheet
for name in list(studentdict.keys()):
    wb.create_sheet(index=1, title=name)
    mainsheet= wb['Sheet']
    link= "=HYPERLINK(\"#\'%s\'!A1\",\"%s\")" % (name,name)
    maxrow=mainsheet.max_row+1
    mainsheet.cell(row=maxrow, column=1).value = link

    sheet = wb[name]
    sheet.cell(row=1, column=1).value = 'Name'
    sheet.cell(row=1, column=2).value = 'Class'
    sheet.cell(row=1, column=3).value = 'Date'
    sheet.cell(row=1, column=4).value = 'Count'
    a=studentdict[name]
    urllist= list(a.keys())
    for url in urllist:
    
        #add class name here
        b=a[url]
        c=b['datecount']
        for eachdate in list(c.keys()):
            maxrow=sheet.max_row +1
            sheet.cell(row=maxrow, column=1).value = name
            sheet.cell(row=maxrow, column=2).value = b['coursename']

            dttm = datetime.datetime.strptime(eachdate, "%m/%d/%Y").date()
            
            sheet.cell(row=maxrow, column=3).value = dttm
            
            sheet.cell(row=maxrow, column=4).value = c[eachdate]

            #column chart
    print(name)
    newname=name.replace(', ','')
    
    newname=newname.replace(' ','')
    
    newname=newname.replace('-','')
    
    newname=newname.replace("'",'')
    newname=newname.replace("(",'')
    newname=newname.replace(")",'')
    print(newname)
    if sheet.max_row>1:
        tab = Table(displayName=str(newname), ref='A1:D%s' % sheet.max_row)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        sheet.add_table(tab)

    if sheet.max_row>2:
##        chart= BarChart()
##        chart.type = "col"
##        chart.style = 10
##        chart.title = name
####        chart.x_axis = DateAxis()
####        chart.x_axis.majorTimeUnit = "days"
##        chart.x_axis.number_format ='mm/dd/yyyy'
##        data = Reference(sheet, min_col=4, min_row=2, max_row=sheet.max_row,)
##        title = Reference(sheet, min_col=3, min_row=2, max_row=sheet.max_row)
##        chart.add_data(data)
##        chart.set_categories(title)
##        sheet.add_chart(chart, "A10")

        chart = BarChart()
        chart.title = name
        chart.style = 10
        chart.x_axis.title = 'Date'
        chart.y_axis.title = 'Count'
        chart.y_axis.crossAx = 500
        chart.x_axis = DateAxis(crossAx=100)
        chart.x_axis.number_format ='yyyy/mm/dd'
        chart.x_axis.majorTimeUnit = "days"
        data = Reference(sheet, min_col=4, min_row=1, max_row=sheet.max_row)
        chart.add_data(data, titles_from_data=True)
        dates = Reference(sheet, min_col=3, min_row=2, max_row=sheet.max_row)
        chart.set_categories(dates)
        sheet.add_chart(chart, "G9")

    sheet['G2'].value ='SUM(D:D)/NETWORKDAYS(MIN(C:C),TODAY(),DATEVALUE({"2017/09/04","2017/09/21","2017/10/19","2017/10/20","2017/11/10","2017/11/22","2017/11/23","2017/11/24","2017/12/22","2017/12/25","2017/12/26","2017/12/27","2017/12/28","2018/12/29","2018/01/01","2018/01/02","2018/01/03","2018/01/04","2018/01/08","2018/01/15","2018/02/19","2018/03/23","2018/03/26","2018/03/27","2018/03/28","2018/03/29","2018/03/30","2018/05/25","2018/05/28","2018/09/03","2018/09/10","2018/09/19","2018/10/19","2018/11/06","2018/11/12","2018/11/22","2018/11/21","2018/11/23","2018/12/24","2018/12/25","2018/12/26","2018/12/27","2018/12/28","2018/12/31","2019/01/01","2019/01/02","2019/01/03","2019/01/04","2019/01/07","2019/01/21","2019/02/18","2019/03/22","2019/03/28","2019/03/27","2019/03/25","2019/03/26","2019/03/27","2019/03/29","2019/04/19","2019/05/27","2018/06/06","2018/06/07","2018/06/08","2018/06/09","2018/06/10","2018/06/11","2018/06/12","2018/06/13","2018/06/14","2018/06/15","2018/06/16","2018/06/17","2018/06/18","2018/06/19","2018/06/20","2018/06/21","2018/06/22","2018/06/23","2018/06/24","2018/06/25","2018/06/26","2018/06/27","2018/06/28","2018/06/29","2018/06/30","2018/07/01","2018/07/02","2018/07/03","2018/07/04","2018/07/05","2018/07/06","2018/07/07","2018/07/08","2018/07/09","2018/07/10","2018/07/11","2018/07/12","2018/07/13","2018/07/14","2018/07/15","2018/07/16","2018/07/17","2018/07/18","2018/07/19","2018/07/20","2018/07/21","2018/07/22","2018/07/23","2018/07/24","2018/07/25","2018/07/26","2018/07/27","2018/07/28","2018/07/29","2018/07/30","2018/07/31","2018/08/01","2018/08/02","2018/08/03","2018/08/04","2018/08/05","2018/08/06","2018/08/07","2018/08/08","2018/08/09","2018/08/10","2018/08/11","2018/08/12","2018/08/13","2018/08/14","2018/08/15","2019/06/05","2019/06/06","2019/06/07","2019/06/08","2019/06/09","2019/06/10","2019/06/11","2019/06/12","2019/06/13","2019/06/14","2019/06/15","2019/06/16","2019/06/17","2019/06/18","2019/06/19","2019/06/20","2019/06/21","2019/06/22","2019/06/23","2019/06/24","2019/06/25","2019/06/26","2019/06/27","2019/06/28","2019/06/29","2019/06/30","2019/07/01","2019/07/02","2019/07/03","2019/07/04","2019/07/05","2019/07/06","2019/07/07","2019/07/08","2019/07/09","2019/07/10","2019/07/11","2019/07/12","2019/07/13","2019/07/14","2019/07/15","2019/07/16","2019/07/17","2019/07/18","2019/07/19","2019/07/20","2019/07/21","2019/07/22","2019/07/23","2019/07/24","2019/07/25","2019/07/26","2019/07/27","2019/07/28","2019/07/29","2019/07/30","2019/07/31","2019/08/01","2019/08/02","2019/08/03","2019/08/04","2019/08/05","2019/08/06","2019/08/07","2019/08/08","2019/08/09","2019/08/10","2019/08/11","2019/08/12","2019/08/13","2019/08/14","2019/08/15","2019/08/16","2019/08/17","2019/08/18","2019/08/19","2019/08/20"}))'
    sheet['F1'].value = 'Active Rate'
    sheet['F2'].value ='SUM(D:D)/SUMPRODUCT(1/COUNTIF(%s[Date],%s[Date]))' % (newname, newname)
    sheet['G1'].value = 'Overall Rate'
    sheet['H1'].value = 'Closing Speed'
    sheet['H2'].value = '=72/G2'
    sheet['H3'].value = 'Days'

    sheet['I3'].value ='Days'
    sheet['M1'].value ='Credits Left'
    sheet['M2'].value ='24'

    sheet['N1'].value ='School Years to Graduation'
    sheet['N2'].value ='=(H2*2*M2)/180'

 

    sheet['Q1'].value ='School Years to Graduation with 1 more'
    sheet['Q2'].value ='=(((72/(G2+1))*2*M2)/180)'


wb.save('C:\\Programs\\Activity\\StudentActivityData.xlsx')


